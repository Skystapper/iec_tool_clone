

# TEST 2: Does file exist?
import os
print(os.path.exists("data/SDLC.xlsx"))

# TEST 3: What's in Excel?
from openpyxl import load_workbook
wb = load_workbook("data/SDLC.xlsx")
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
    print(row)

# TEST 4: Search works?
from app.services.excel_service import get_requirement_data
result = get_requirement_data("SG", "1.a")
print(f"Result: {result}")
