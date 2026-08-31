"""
Reads the IEC 62443 requirements database (data/SDLC.xlsx).

The ABBY IEC 62443 agent groups certain cross-cutting requirements with a
category run even though they live under a different category in the
spreadsheet. In particular, an SG assessment also includes the legal
disclaimer requirement ABB-SDLC-10. We mirror that grouping here so our output
matches the agent's native report.
"""
from functools import lru_cache
from typing import Dict, List

from openpyxl import load_workbook

from app.core.paths import SDLC_XLSX

# Requirements that should be appended when assessing a given category.
# Key = category being run; value = list of (category, id) pairs to include.
CROSS_CATEGORY = {
    "SG": [("ABB", "ABB-SDLC-10")],
}


@lru_cache(maxsize=1)
def _load_rows() -> List[dict]:
    """Load every requirement row once and cache it."""
    wb = load_workbook(SDLC_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None or row[1] is None:
            continue
        rows.append({
            "category": str(row[0]).strip().upper(),
            "id": str(row[1]).strip(),
            "requirement": row[2],
            "rationale": row[3],
            "guidance": row[4],
        })
    wb.close()
    return rows


def _make(row: dict) -> dict:
    return {
        "id": row["id"],
        "full_id": f'{row["category"]}-{row["id"]}',
        "requirement": row["requirement"],
        "rationale": row["rationale"],
        "guidance": row["guidance"],
    }


def get_requirement_data(category: str, req_id: str):
    """Get a single requirement, e.g. get_requirement_data('SG', '1.a')."""
    cat = str(category).strip().upper()
    rid = str(req_id).strip()
    for row in _load_rows():
        if row["category"] == cat and row["id"] == rid:
            return {
                "requirement": row["requirement"],
                "rationale": row["rationale"],
                "guidance": row["guidance"],
            }
    return None


def get_all_requirements_for_category(category: str) -> List[dict]:
    """
    Return all requirements for a category, including any cross-cutting
    requirements that the agent groups with it (e.g. ABB-SDLC-10 for SG).
    """
    cat = str(category).strip().upper()
    out: List[dict] = []
    for row in _load_rows():
        if row["category"] == cat:
            out.append(_make(row))

    for xcat, xid in CROSS_CATEGORY.get(cat, []):
        for row in _load_rows():
            if row["category"] == xcat and row["id"] == xid:
                out.append(_make(row))

    return out
