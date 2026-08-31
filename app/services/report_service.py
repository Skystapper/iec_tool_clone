"""
Generates the styled Excel compliance report.

The visual style matches the reports produced by ABBY's IEC 62443 agent:
bold white header on the ABB red, wrapped text, thin borders, frozen header,
auto-filter, sensible column widths, and color-coded Status cells (green /
amber / red).
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.paths import output_report_path

# ABB brand red used in the agent's reference report.
HEADER_FILL = PatternFill("solid", fgColor="FF000F")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

STATUS_FILLS = {
    "fully met": PatternFill("solid", fgColor="C6EFCE"),        # green
    "partially met": PatternFill("solid", fgColor="FFEB9C"),    # amber
    "not met": PatternFill("solid", fgColor="FFC7CE"),          # red
    "not addressed": PatternFill("solid", fgColor="D9D9D9"),    # grey
    "not assessed": PatternFill("solid", fgColor="D9D9D9"),      # grey
    "error": PatternFill("solid", fgColor="FFC7CE"),            # red
}
STATUS_FONTS = {
    "fully met": Font(color="006100", bold=True),
    "partially met": Font(color="9C5700", bold=True),
    "not met": Font(color="9C0006", bold=True),
    "not addressed": Font(color="595959", bold=True),
    "not assessed": Font(color="595959", bold=True),
    "error": Font(color="9C0006", bold=True),
}

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER_TOP = Alignment(wrap_text=True, vertical="top", horizontal="center")

HEADERS = [
    "File Name", "Category", "ID", "Requirement", "Rationale",
    "Guidance", "Status", "Explanation", "Evidence", "Recommendations",
]

# Column widths (character units) tuned to match the ABB report's readability.
COL_WIDTHS = {
    "A": 22,  # File Name
    "B": 10,  # Category
    "C": 12,  # ID
    "D": 50,  # Requirement
    "E": 45,  # Rationale
    "F": 45,  # Guidance
    "G": 16,  # Status
    "H": 55,  # Explanation
    "I": 55,  # Evidence
    "J": 55,  # Recommendations
}


def generate_excel(results, session_id) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Report"

    # --- Header row ---------------------------------------------------------
    ws.append(HEADERS)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_TOP
        cell.border = BORDER

    # --- Data rows ----------------------------------------------------------
    for r in results:
        ws.append([
            r.get("file_name", ""),
            r.get("category", ""),
            r.get("id", ""),
            r.get("requirement", ""),
            r.get("rationale", ""),
            r.get("guidance", ""),
            r.get("status", ""),
            r.get("explanation", ""),
            r.get("evidence", ""),
            r.get("recommendations", ""),
        ])

    # Apply borders/wrap to every data cell and colour-code the Status column.
    status_col = HEADERS.index("Status") + 1  # 1-based
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.border = BORDER
            cell.alignment = CENTER_TOP if cell.column == status_col else WRAP_TOP
        status_cell = row[status_col - 1]
        key = str(status_cell.value or "").strip().lower()
        if key in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[key]
            status_cell.font = STATUS_FONTS[key]

    # --- Column widths, freeze, filter -------------------------------------
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    ws.row_dimensions[1].height = 30

    file_path = output_report_path(str(session_id))
    wb.save(file_path)
    return str(file_path)
